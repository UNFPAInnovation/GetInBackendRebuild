import datetime
import traceback
from tempfile import TemporaryFile

import pytz
import xlrd
from django.utils import timezone

from GetInBackendRebuild.settings import SHEET_FILES_FOLDER
from app.airtime_dispatcher import AirtimeModule
from app.models import District, County, SubCounty, Parish, Village, User, Girl, Appointment, FollowUp, HealthFacility, \
    Delivery, MSIService, Region
from app.utils.utilities import add_months, unique
from openpyxl import Workbook, load_workbook
from xlwt import Workbook as WorkbookCreation
from app.utils.constants import *
from django.db.models.expressions import Q
from django.db.models import Case, When, IntegerField, Count


def create_odk_sheets(location):
    """
    :param location: uri of the excel file to extract from. Usually excel files are placed in the sheets folder
    Creates odk sheets for each district on the second sheet of the odk sheet template
    Sheet Fields: Region, District, County, SubCounty, Parish, Village
    """
    wb = xlrd.open_workbook(location)
    sheet = wb.sheet_by_index(0)

    odk_wb = load_workbook(SHEET_FILES_FOLDER + 'odksample.xlsx')
    odk_sheet = odk_wb.get_sheet_by_name('choices')
    settings_sheet = odk_wb.get_sheet_by_name('settings')
    counties = []
    subcounties = []
    parishes = []
    villages = []
    district_name = ""

    for row_number in range(0, 78999):
        try:
            row_data = sheet.row_values(row_number)
        except Exception as e:
            print(e)
            break

        print(row_data)
        district_value = row_data[1]
        county_value = row_data[2]
        sub_county_value = row_data[3]
        parish_value = row_data[4]
        village_value = row_data[5]

        # skip first row with headers
        if str(district_value).lower() == 'district':
            continue

        if district_name != district_value and district_name:
            for county in unique(counties):
                odk_sheet.append(county)

            for subcounty in unique(subcounties):
                odk_sheet.append(subcounty)

            for parish in unique(parishes):
                odk_sheet.append(parish)

            for village in unique(villages):
                odk_sheet.append(village)

            settings_sheet.append(
                ['GetInMapGirl' + district_name + '1_chew', 'GetInMapGirl' + district_name + '1_chew', '', '', '','yes'])
            odk_wb.save(SHEET_FILES_FOLDER + 'GetInMapGirl' + district_name + '1_chew.xls')
            counties = []
            subcounties = []
            parishes = []
            villages = []
            odk_wb = load_workbook(SHEET_FILES_FOLDER + 'odksample.xlsx')
            odk_sheet = odk_wb.get_sheet_by_name('choices')
            settings_sheet = odk_wb.get_sheet_by_name('settings')
            district_name = district_value

            # stop on reaching end of document
            if not district_value:
                break
        else:
            counties.append(['countys', county_value, county_value])
            subcounties.append(['subcountys', sub_county_value, sub_county_value, county_value])
            parishes.append(['parishs', parish_value, parish_value, county_value, sub_county_value])
            villages.append(['villages', village_value, village_value, county_value, sub_county_value, parish_value])
            district_name = district_value


def extract_excel_org_unit_data(location):
    """
    :param location: uri of the excel file to extract from. Usually excel files are placed in the sheets folder
    Creates org units from an excel sheet provided
    Sheet Fields: Region, District, County, SubCounty, Parish, Village
    Note: Some districts don't have county. In that case the district name is used as a county
    """
    wb = xlrd.open_workbook(location)
    sheet = wb.sheet_by_index(0)

    for row_number in range(0, sheet.utter_max_rows):
        try:
            row_data = sheet.row_values(row_number)
        except Exception as e:
            print(e)
            break

        print(row_data)
        region_value = row_data[0]
        district_value = row_data[1]
        county_value = row_data[2]
        sub_county_value = row_data[3]
        parish_value = row_data[4]
        village_value = row_data[5]

        # stop on reaching end of document
        if not district_value:
            break

        # skip first row with headers
        if str(district_value).lower() == 'district':
            continue

        region = Region.objects.get_or_create(name=get_region_name(region_value))
        district = District.objects.get_or_create(name=district_value, region=region[0])
        county = County.objects.get_or_create(name=county_value, district=district[0])
        sub_county = SubCounty.objects.get_or_create(name=sub_county_value, county=county[0])
        parish = Parish.objects.get_or_create(name=parish_value, sub_county=sub_county[0])
        village = Village.objects.get_or_create(name=village_value, parish=parish[0])


def get_region_name(region_id):
    region_id = int(float(region_id))
    if region_id == 1:
        return "Central"
    elif region_id == 2:
        return "Eastern"
    elif region_id == 3:
        return "Northern"
    elif region_id == 4:
        return "Western"
    else:
        return "West Nile"


def extract_excel_user_data(location, district_name):
    """
    :param location: uri of the excel file of users to extract from. Usually excel files are placed in the sheets folder
    :param district_name: district where these users will be created
    Creates user accounts from an excel sheet provided
    Sheet Fields: FirstName, LastName, Personal PhoneNumber, GetIn PhoneNumber, Role, Gender, HealthCenter, SubCounty
    """
    wb = xlrd.open_workbook(location)
    sheet = wb.sheet_by_index(0)

    for row_number in range(0, 1000):
        try:
            row_data = sheet.row_values(row_number)
        except Exception as e:
            print(e)
            break

        first_name_value = row_data[0]
        last_name_value = row_data[1]
        getin_number_value = int(row_data[3])
        role = row_data[4]
        gender = row_data[5]
        health_facility = row_data[6]
        sub_county = row_data[7]

        if not first_name_value:
            break

        try:
            sub_county = SubCounty.objects.get(name__contains=sub_county)
            village = sub_county.parish_set.first().village_set.first()
        except Exception as e:
            print(e)
            sub_county = District.objects.get(name__icontains=district_name).county_set.last() \
                .subcounty_set.last()
            village = sub_county.parish_set.last().village_set.last()

        try:
            health_facility, _ = HealthFacility.objects.get_or_create(name=health_facility, sub_county=sub_county,
                                                                      facility_level=len(
                                                                          health_facility.split("HC")[1]))
        except Exception as e:
            print(e)
            health_facility, _ = HealthFacility.objects.get_or_create(name=health_facility, sub_county=sub_county,
                                                                      facility_level='0')

        try:
            getin_number_value = "0" + str(int(getin_number_value))
            user = User(first_name=first_name_value.strip(), village=village, last_name=last_name_value.strip(),
                        username=str(first_name_value.strip()).lower()[0] + str(
                            last_name_value.strip()).lower().replace(" ", ""),
                        email=first_name_value + last_name_value + "@getinmobile.org", phone=getin_number_value.strip(),
                        health_facility=health_facility,
                        gender=GENDER_FEMALE if gender.lower().find("f") > -1 else GENDER_MALE,
                        role=USER_TYPE_CHEW if str(role).lower().find("vht") > -1 else str(role).lower())
            user.set_password(getin_number_value)
            user.save()
        except Exception:
            print(traceback.print_exc())


def generate_monthly_system_stats(location=SHEET_FILES_FOLDER + 'GetIN Traceability Form ' + str(
    timezone.now().strftime("%m-%d-%Y, %H:%M")) + '.xls', end_date=None):
    """
    :param location: name of the excel file containing the output
    :param end_date: date up to which the metrics will be generated
    Creates user performance excel sheet
    Sheet Fields: NAME OF HEALTHWORKER, PHONE NO., ROLE, DISTRICT, MONTH, YEAR, NO. OF GIRLS MAPPED, ATTENDED, MISSED, EXPECTED, FOLLOW UPS
    """
    book = WorkbookCreation()
    sheet1 = book.add_sheet('Sheet1')

    data = [
        ["NAME OF HEALTHWORKER"], ["PHONE NO."], ["ROLE"], ["DISTRICT"], ["MONTH"], ["YEAR"],
        ["NO. OF GIRLS MAPPED"], ["ATTENDED"], ["MISSED"], ["EXPECTED"], ["FOLLOW UPS"]
    ]

    for c_index, columns in enumerate(data):
        for r_index, row_item in enumerate(columns):
            sheet1.write(r_index, c_index, row_item)

    book.save(location)
    book.save(TemporaryFile())

    convert_xls_to_xlsx(location)

    for district in District.objects.all():
        created_at = timezone.datetime(2019, 10, 1).replace(tzinfo=pytz.utc)
        if not end_date:
            end_date = timezone.now()

        while created_at < end_date.replace(tzinfo=pytz.utc):
            for user in User.objects.filter(district=district):
                if user.role in [USER_TYPE_DHO, USER_TYPE_AMBULANCE, USER_TYPE_DEVELOPER]:
                    continue
                girls = Girl.objects.filter(Q(created_at__gte=created_at) &
                                            Q(created_at__lte=add_months(created_at, 1)
                                              .replace(tzinfo=pytz.utc)) & Q(user=user)).count()
                wb = load_workbook(location.replace('xls', 'xlsx'))
                sheet = wb['Sheet1']

                appointment_data = {
                    "attended": 0,
                    "expected": 0,
                    "missed": 0
                }

                if user.role == USER_TYPE_MIDWIFE:
                    appointment_data = Appointment.objects.filter(Q(user=user) & Q(created_at__gte=created_at) &
                                                                  Q(created_at__lte=add_months(created_at, 1)
                                                                    .replace(tzinfo=pytz.utc))).aggregate(
                        attended=Count(Case(
                            When(Q(status=ATTENDED), then=1),
                            output_field=IntegerField(),
                        )),
                        expected=Count(Case(
                            When(Q(status=EXPECTED), then=1),
                            output_field=IntegerField(),
                        )),
                        missed=Count(Case(
                            When(Q(status=MISSED), then=1),
                            output_field=IntegerField(),
                        ))
                    )

                followups = FollowUp.objects.filter(Q(user=user) & Q(created_at__gte=created_at) &
                                                    Q(created_at__lte=add_months(created_at, 1).replace(
                                                        tzinfo=pytz.utc))).count()

                sheet.append([user.first_name + " " + user.last_name, user.phone, user.role, district.name,
                              created_at.strftime("%B"), created_at.strftime("%Y"), girls, appointment_data['attended'],
                              appointment_data['expected'],
                              appointment_data['missed'], followups])
                wb.save(location.replace('xls', 'xlsx'))
            created_at = add_months(created_at, 1).replace(tzinfo=pytz.utc)


def convert_xls_to_xlsx(src_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)
    book_xlsx.save(src_file_path.replace('xls', 'xlsx'))


def extract_excel_user_data_for_airtime_dispatchment(location):
    """
    :param location: excel file containing the phone numbers
    :param amount: the amount of airtime that will be loaded on the users' phones
    Sends airtime to users GetIN phone numbers
    """
    wb = xlrd.open_workbook(location)
    sheet = wb.sheet_by_index(0)
    phone_numbers = []

    for row_number in range(0, sheet.utter_max_rows):
        try:
            row_data = sheet.row_values(row_number)
            phone_numbers.append("+256" + str(int(row_data[0])))
        except Exception as e:
            print(e)
            break
    return list(set(phone_numbers))


def send_airtime_to_users(location, amount):
    phone_numbers = extract_excel_user_data_for_airtime_dispatchment(location)
    airtime = AirtimeModule()
    airtime.send_airtime(phone_numbers, amount)


def generate_user_credential_sheet(district_name):
    for user in User.objects.filter(district__name=district_name):
        girls = Girl.objects.filter(user=user).count()
        location = SHEET_FILES_FOLDER + "GetIN User Credentials.xlsx"
        wb = load_workbook(location)
        sheet = wb['Sheet1']
        sheet.append([user.first_name + " " + user.last_name, user.username, user.phone, user.role, girls])
        wb.save(location)


def generate_overall_stats(district, from_date=datetime.datetime(2019, 11, 1), to_date=datetime.datetime(2060, 1, 1)):
    result = dict()

    result["Mapped girls"] = Girl.objects.filter(
        Q(created_at__gte=from_date) & Q(created_at__lte=to_date) & Q(user__district__name=district)).count()
    result["ANC visits"] = Appointment.objects.filter(
        Q(created_at__gte=from_date) & Q(created_at__lte=to_date) & Q(user__district__name=district)).count()
    result["Follow ups"] = FollowUp.objects.filter(
        Q(created_at__gte=from_date) & Q(created_at__lte=to_date) & Q(user__district__name=district)).count()
    result["Deliveries"] = Delivery.objects.filter(
        Q(created_at__gte=from_date) & Q(created_at__lte=to_date) & Q(user__district__name=district)).count()
    result["Voucher services redeemed"] = MSIService.objects.filter(
        Q(created_at__gte=from_date) & Q(created_at__lte=to_date) & Q(girl__user__district__name=district)).count()
    result["Voucher cards created"] = Girl.objects.filter(
        Q(created_at__gte=from_date) & Q(created_at__lte=to_date) & Q(user__district__name=district) &
        Q(voucher_number__contains="-")).count()
    return result
